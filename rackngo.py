#!/usr/bin/python3


from openpyxl import load_workbook
import warnings
import re
import requests
import os
import json

warnings.simplefilter("ignore")

class StaffAPI():
    def __init__(self):
        self.url = "https://api.packet.net/"
        self.headers = {"X-Auth-Token":"{api_token}".format(api_token=os.environ['PACKET_API_AUTH']), "X-Consumer-Token":"{consumer_token}".format(consumer_token=os.environ['PACKET_API_CONSUMER_TOKEN']), "X-Packet-Staff": "true"}
    def get_providers(self,query_string,query_string_value,key):
        endpoint = "staff/providers"
        query_string = query_string+query_string_value
        try:
            response = requests.get(self.url+endpoint+query_string, headers=self.headers)
        except ConnectionError as error:
            print(error)
        return response.json()['providers'][0][key]
    def get_manufacturer_id (self, manufacturer_name) -> str:
        return self.get_providers(query_string="?name=",query_string_value=manufacturer_name,key="id")
    def get_server_racks(self,query_string,query_string_value,key,pod_code):
        endpoint = "staff/server-racks"
        query_string = query_string+query_string_value
        try:
            response = requests.get(self.url+endpoint+query_string, headers=self.headers)
            server_count = response.json()['meta']['total_count']
            if server_count > 1:
                for rack in response.json()['server_racks']:
                    if rack['pod_code'] == pod_code:
                        server_rack_id = rack[key]
            else:
                server_rack_id = response.json()['server_racks'][0][key]   
        except ConnectionError as error:
            return error
        return server_rack_id
    def get_server_rack_id (self,rack_name,pod_code ) -> str:
        return self.get_server_racks(query_string="?name=", query_string_value=rack_name,key="id",pod_code=pod_code)
    def get_hardware(self,query_string,query_string_value):
        endpoint = "staff/hardware"
        query_string = query_string+query_string_value
        try:
            response = requests.get(self.url+endpoint+query_string, headers=self.headers)
        except ConnectionError as error:
            return error
        return response.json()
    def create_hardware(self,server_rack_id,payload):
        endpoint="staff/server-racks/{rack_uuid}/hardware".format(rack_uuid=server_rack_id)
        self.headers.update({"Content-Type": "application/json"})
        try:
            response = requests.post(self.url+endpoint, data=json.dumps(payload), headers=self.headers)
        except ConnectionError as error:
            return error
        return response.json()
    def delete_hardware(self,hardware_id):
        endpoint="staff/hardware/bulk"
        self.headers.update({"Content-Type": "application/json"})
        try:
            response = requests.delete(self.url+endpoint, data=json.dumps({'hardware_ids': ['{hardware_id}'.format(hardware_id=hardware_id)]}), headers=self.headers)
        except ConnectionError as error:
            return error
        return response.json()
    def rack_hardware(self,slot,hardware_id):
        endpoint="/staff/hardware/{new_hardware_uuid}/rackmount".format(new_hardware_uuid=hardware_id)
        self.headers.update({"Content-Type": "application/json"})
        try:
            response = requests.post(self.url+endpoint, data=json.dumps({"top_slot":slot}), headers=self.headers)
        except ConnectionError as error:
            return error
        return response.json()
    def unrack_hardware(self,hardware_id):
        endpoint="/staff/hardware/{new_hardware_uuid}/unrackmount".format(new_hardware_uuid=hardware_id)
        self.headers.update({"Content-Type": "application/json"})
        try:
            response = requests.post(self.url+endpoint, headers=self.headers)
        except ConnectionError as error:
            return error
        return response.json()

class RackDevice(object):
    def __init__(self, name, type,uspaces,model_number,asset_type, manufacturer_id, conn: StaffAPI()):
        self.name = name
        self.type = type
        self.asset_type = asset_type
        self.uspaces = uspaces
        self.model_number = model_number
        self.description = asset_type
        self.manufacturer_id = manufacturer_id
        self.conn = conn
    def get_payload(self):
        payload = {}
        payload.update({"name":self.name})
        payload.update({"type":self.type})
        payload.update({"asset_type":self.asset_type})
        payload.update({"description":self.description})
        payload.update({"model_number":self.model_number})
        payload.update({"manufacturer_id":self.manufacturer_id})
        payload.update({"u_spaces":self.uspaces})
        return payload
    def exists (self) -> bool:
        if self.conn.get_hardware(query_string="?name=",query_string_value=self.name)['meta']['total_count'] > 0:
            return True
        else:
            return False    
    def get_hardware_id (self):
        if self.exists() is True:
            return self.conn.get_hardware(query_string="?name=",query_string_value=self.name)['hardware'][0]['id']
        else:
            raise Exception("No hardware_id for {device}".format(device=self.name))
    def add_hardware_to_rack (self):
        self.conn.create_hardware(self.server_rack_id,RackDevice.get_payload(self))
        print("Creating {device} complete".format(device=device.name))
        self.conn.rack_hardware(self.low_slot, RackDevice.get_hardware_id(self))
        print("Racking {device} complete".format(device=device.name))
    def delete_hardware_from_rack (self):
        self.conn.unrack_hardware(RackDevice.get_hardware_id(self))
        print("UnRacking {device} complete".format(device=device.name))
        self.conn.delete_hardware(RackDevice.get_hardware_id(self))
        print("Deleting {device} complete".format(device=device.name))

def get_devices_from_racks (file, conn: StaffAPI()) -> list:
    racks = []
    wb = load_workbook(file)

    for sheet in wb.sheetnames:
        #if "LA4 (LA)_TEST_EQX" in sheet:
        site = sheet.split()[0].lower()
        sheet = wb[sheet]
        merged_cells = sheet.merged_cells
        for cols in sheet.iter_cols():
            for cell in cols:
                if cell.value is True:
                    rack_enabled = True
                if cell.value is False:
                    rack_enabled = False
                if type(cell.value) is str:
                    if re.search(r'^p\d\d',cell.value):
                        pod_code = cell.value
                    elif re.search(r'^RK\d+',cell.value):
                            rack_name = cell.value.lower()+"."+site+".packet.net"
                    elif "Panel" in cell.value:
                        if cell.coordinate in merged_cells:
                            for cellranges in merged_cells:
                                if cell.coordinate in cellranges:
                                    uspaces = (cellranges.max_row - cellranges.min_row) + 1
                                    low_slot = sheet.cell(row=cellranges.max_row, column=1).value
                        else:
                            uspaces = 1
                            low_slot = sheet.cell(row=cell.row, column=1).value
                        device_type = "Asset"
                        asset_type = 'Panel'
                        pod_code='p01'
                        manufacturer_name = "Generic"
                        model_number = "Panel 1"
                        new_device = RackDevice(name=asset_type.lower()+".u"+str(low_slot)+"."+rack_name, type=device_type,asset_type=asset_type,model_number=model_number,manufacturer_id=conn.get_manufacturer_id(manufacturer_name=manufacturer_name),uspaces=uspaces, conn=conn)
                        new_device.pod_code = pod_code
                        new_device.rack_name = rack_name
                        new_device.low_slot = int(low_slot)
                        new_device.rack_enabled = rack_enabled if type(rack_enabled) is bool else False
                        new_device.server_rack_id = conn.get_server_rack_id(rack_name=rack_name,pod_code=pod_code)
                        racks.append(new_device)
    return racks

if __name__ == '__main__':
    staffapi = StaffAPI()
    try:
        excel_file = os.environ['PACKET_RACK_SHEET']
        os.path.exists(excel_file)
    except:
        excel_file = input("Excel File to Process: ")
    for device in get_devices_from_racks(excel_file, staffapi):
        if device.rack_enabled is True:
            if device.exists() is not True:
                device.add_hardware_to_rack()
            # else:
            #     device.delete_hardware_from_rack()
